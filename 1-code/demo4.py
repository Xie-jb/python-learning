#Excel练习
# 4. 把筛选结果保存成新的 Excel。

from openpyxl import load_workbook, Workbook

wb = load_workbook(r"C:\codex\2026.5.31-python-learning-demo1\2-file\orders.xlsx")
sheet = wb.active
wbNew = Workbook()
sheet2 = wbNew.active
rowNew = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    for str in row:
        if str == "未付款":
            sheet2.append(row)
            print(row)
wbNew.save(r"C:\codex\2026.5.31-python-learning-demo1\2-file\not_pay.xlsx")