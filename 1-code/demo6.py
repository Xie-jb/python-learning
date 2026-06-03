# ## 练习任务

# 1. 准备一个订单 Excel 示例文件。
# 2. 编写读取函数。
# 3. 编写筛选函数。
# 4. 编写统计函数。
# 5. 导出结果 Excel。
# 6. 测试不同订单状态是否正确。

from openpyxl import load_workbook, Workbook
wb = load_workbook(r"C:\codex\2026.5.31-python-learning-demo1\2-file\day07_orders_practice.xlsx")
file = wb.active

wbOut1 = Workbook()
fileOut1 = wbOut1.active

wbOut2 = Workbook()
fileOut2 = wbOut2.active

regions = set()
region_order = {}
states = set()
order_state = {}
for row in file.iter_rows(min_row=2,values_only=True):
    region = row[3]
    regions.add(region)
    for str in row:
        if str == "已付款":
            fileOut1.append(row)
        elif str == "未付款":
            fileOut2.append(row)

for row in file.iter_rows(min_row=2,values_only=True):
    region = row[3]
    if region not in region_order:
        region_order[region] = []
    region_order[region].append(row)

for row in file.iter_rows(min_row=2,values_only=True):
    state = row[4]
    states.add(state)
    if state not in order_state:
        order_state[state] = []
    order_state[state].append(row)

print(region_order["华南"])
print(states)

wbOut1.save(r"C:\codex\2026.5.31-python-learning-demo1\2-file\pay.xlsx")
wbOut2.save(r"C:\codex\2026.5.31-python-learning-demo1\2-file\not_pay.xlsx")

