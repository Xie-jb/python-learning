# Excel练习
# 1. 读取订单 Excel。
# 2. 统计订单总金额。
# 3. 筛选“未付款”订单。
# 4. 把筛选结果保存成新的 Excel。



from openpyxl import load_workbook, Workbook

wb = load_workbook(r"C:\codex\2026.5.31-python-learning-demo1\2-file\orders.xlsx")
sheet = wb.active

for row in sheet.iter_rows(min_row=2, values_only=True):
    print(row)
